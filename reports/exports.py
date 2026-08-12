import io

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _flat_rows(rows):
    """Convert row dicts (cells/actions) to plain lists of cell values."""
    return [list(r.get("cells", [])) for r in rows]


def export_xlsx(title, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in _flat_rows(rows):
        ws.append([str(c) for c in row])

    ws.freeze_panes = "A2"
    for i, _ in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    wb.save(response)
    return response


def export_pdf(title, columns, rows):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=title,
    )

    story = [
        Paragraph(f"<b>Srinivasa Technology</b>", report_style(size=10)),
        Paragraph(title, report_style(size=14)),
        Spacer(1, 8 * mm),
    ]

    data = [columns] + _flat_rows(rows)
    if len(data) > 1:
        col_widths = [None] * len(columns)
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF1F8")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
    else:
        story.append(Paragraph("<i>No data</i>", report_style(size=10)))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title}.pdf"'
    return response


def report_style(size=10):
    from reportlab.lib.styles import ParagraphStyle

    return ParagraphStyle(
        name=f"rpt{size}",
        fontName="Helvetica-Bold",
        fontSize=size,
        spaceAfter=4,
        textColor=colors.HexColor("#1F4E79"),
    )
