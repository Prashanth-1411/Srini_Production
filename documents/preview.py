import html
from datetime import date, datetime, time

MAX_ROWS = 300
MAX_COLS = 30


def _fmt(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(sep=" " if isinstance(value, datetime) else "")
    return value


def _sheet_to_html(name, rows):
    total_rows = len(rows)
    total_cols = max((len(r) for r in rows), default=0)
    nrows = min(total_rows, MAX_ROWS)
    ncols = min(total_cols, MAX_COLS)

    head = []
    body = []
    for r in range(nrows):
        cells = []
        for c in range(ncols):
            raw = rows[r][c] if c < len(rows[r]) else ""
            text = html.escape(str(raw))
            if r == 0:
                cells.append(f"<th>{text}</th>")
            else:
                cells.append(f"<td>{text}</td>")
        row_html = "<tr>" + "".join(cells) + "</tr>"
        (head if r == 0 else body).append(row_html)

    note = ""
    if total_rows > MAX_ROWS or total_cols > MAX_COLS:
        note = (
            f'<div class="alert alert-warning small mb-2">Preview limited to first '
            f"{nrows} rows x {ncols} columns (file has {total_rows} x {total_cols}). "
            f"Download to see the full file.</div>"
        )

    return (
        '<div class="mb-3">'
        f'<div class="fw-semibold small text-muted mb-1">{html.escape(name)}</div>'
        f"{note}"
        '<div class="table-responsive">'
        '<table class="table table-bordered table-sm excel-preview mb-0">'
        f"<thead class=\"table-light\">{''.join(head)}</thead>"
        f"<tbody>{''.join(body)}</tbody>"
        "</table>"
        "</div>"
        "</div>"
    )


def _xlsx_to_html(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sections = []
        for ws in wb.worksheets:
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_ROWS + 1:
                    break
                rows.append([_fmt(v) for v in row])
            sections.append(_sheet_to_html(ws.title, rows))
    finally:
        wb.close()
    return _wrap(sections)


def _xls_to_html(path):
    import xlrd

    wb = xlrd.open_workbook(path, on_demand=True)
    try:
        sections = []
        for sheet in wb.sheets():
            rows = []
            for r in range(min(sheet.nrows, MAX_ROWS + 1)):
                row = []
                for c in range(min(sheet.ncols, MAX_COLS + 1)):
                    cell = sheet.cell(r, c)
                    row.append(_xls_cell(cell, sheet))
                rows.append(row)
            sections.append(_sheet_to_html(sheet.name, rows))
    finally:
        wb.release_resources()
    return _wrap(sections)


def _xls_cell(cell, sheet):
    import xlrd

    ctype = cell.ctype
    if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if ctype == xlrd.XL_CELL_NUMBER:
        value = float(cell.value)
        return str(int(value)) if value.is_integer() else repr(value)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, sheet.book.datemode).isoformat(sep=" ")
        except Exception:
            return str(cell.value)
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    return str(cell.value)


def _wrap(sections):
    if not sections:
        sections = ['<div class="alert alert-info mb-0">This workbook contains no sheets.</div>']
    return '<div class="excel-preview-wrap">' + "".join(sections) + "</div>"


def excel_to_html(path):
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""
    if ext == "xlsx":
        return _xlsx_to_html(path)
    if ext == "xls":
        return _xls_to_html(path)
    raise ValueError("Not an Excel file: %s" % path)
